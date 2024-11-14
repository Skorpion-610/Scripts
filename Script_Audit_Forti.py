import json
import pandas as pd
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference

# Fonction pour ouvrir une fenêtre de dialogue et sélectionner un fichier JSON
def import_json_file():
    root = tk.Tk()
    root.withdraw()  # Cacher la fenêtre principale
    file_path = filedialog.askopenfilename(filetypes=[("JSON files", "*.json")])
    return file_path

# Obtenir le chemin du fichier JSON en utilisant la fonction d'importation
file_path = import_json_file()

# Vérifier si un fichier a été sélectionné
if file_path:
    # Charger les règles depuis le fichier JSON sélectionné
    with open(file_path, 'r') as file:
        rules = json.load(file)

    # Convertir les règles en DataFrame pour une analyse facile
    df = pd.DataFrame(rules)

    # Fonction pour vérifier les règles + Nom des feuilles Excel
    def verify_rules(df):
        findings = {
            'General Checks': [],
            'Source and Destination Checks': [],
            'Service Checks': [],
            'Bytes Transferred Checks': []
            # Autres catégories si nécessaire
        }

        # Vérifications
        for index, rule in df.iterrows():
            rule_id = rule['ID']
            source_ip = rule['Source']
            destination_ip = rule['Destination']
            action = rule['Action']
            service = rule['Service']
            bytes_transferred = rule['Bytes']
            
            # Détails de la règle
            rule_details = {
                'ID': rule_id,
                'Source': source_ip,
                'Destination': destination_ip,
                'Action': action,
                'Service': service,
                'Bytes Transferred': bytes_transferred
            }

            # Vérifier les IPs source et destination pour "0.0.0.0/0"
            if source_ip == '0.0.0.0/0' or destination_ip == '0.0.0.0/0':
                findings['General Checks'].append({'Finding': f"Rule {rule_id}: Wide open IP range (0.0.0.0/0).",'Details': rule_details})

            # Vérifier les IPs source et destination en "all"
            if source_ip.lower() == 'all':
                findings['Source and Destination Checks'].append({'Finding': f"Rule {rule_id}: Source is set to 'all' with action '{action}'.",'Details': rule_details})
            if destination_ip.lower() == 'all':
                findings['Source and Destination Checks'].append({'Finding': f"Rule {rule_id}: Destination is set to 'all' with action '{action}'.",'Details': rule_details})
                
            # Vérifier les services en "all"
            if service.lower() == 'all':
                findings['Service Checks'].append({'Finding': f"Rule {rule_id}: Service is set to 'all' with action '{action}'.",'Details': rule_details})

            # Vérifier l'action
            if action not in ['ACCEPT', 'DENY']:
                findings['General Checks'].append({'Finding': f"Rule {rule_id}: Unrecognized action '{action}'.",'Details': rule_details})
                
            # Vérifier les bytes transférés avec la valeur "0 B"
            if bytes_transferred.strip() == '0 B':
                findings['Bytes Transferred Checks'].append({'Finding': f"Rule {rule_id}: La règle n'est pas utilisée.",'Details': rule_details})

            # Autres vérifications si nécessaire

        return findings

    # Exécuter les vérifications
    findings = verify_rules(df)

    # Fonction pour enregistrer le fichier Excel à l'emplacement choisi par l'utilisateur
    def save_excel_file():
        root = tk.Tk()
        root.withdraw()  # Cacher la fenêtre principale
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            # Créer un fichier Excel avec différentes feuilles pour chaque type de vérification
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, findings_list in findings.items():
                    # Convertir la liste de résultats en DataFrame pour cette feuille
                    df_sheet = pd.DataFrame([{
                        'Rule ID': finding['Details']['ID'],
                        'Finding': finding['Finding'],
                        'From': rule['From'],  # Utiliser From depuis la règle
                        'To': rule['To'],      # Utiliser To depuis la règle
                        'Source': finding['Details']['Source'],
                        'Destination': finding['Details']['Destination'],
                        'Action': finding['Details']['Action'],
                        'Service': finding['Details']['Service'],
                        'Bytes Transferred': finding['Details']['Bytes Transferred']
                    } for finding in findings_list for _, rule in df.iterrows() if finding['Details']['ID'] == rule['ID']])
                    # Écrire ce DataFrame dans la feuille correspondante du fichier Excel
                    df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

            # Charger le fichier Excel pour appliquer des styles (le reste du code reste inchangé)
            workbook = load_workbook(file_path)

            # Style pour les en-têtes
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))

            for sheet_name in findings.keys():
                worksheet = workbook[sheet_name]
                for col in range(1, worksheet.max_column + 1):
                    col_letter = get_column_letter(col)
                    worksheet.column_dimensions[col_letter].width = 15  # Ajuster la largeur des colonnes
                    cell = worksheet[f"{col_letter}1"]
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border

                # Ajouter des bordures aux cellules
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border

                # Mise en surbrillance conditionnelle pour les anomalies
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    finding_cell = row[1]  # Colonne 'Finding'
                    if "Wide open IP range" in finding_cell.value or "set to 'all'" in finding_cell.value or "Unrecognized action" in finding_cell.value or "n'est pas utilisée" in finding_cell.value:
                        for cell in row:
                            cell.fill = red_fill

                # Geler les volets (conserver la première ligne visible)
                worksheet.freeze_panes = worksheet['A2']

            # Ajouter la feuille récapitulative
            summary_sheet = workbook.create_sheet(title="Summary", index=0)

            # Compter le nombre de règles en erreur par feuille
            summary_data = {"Sheet Name": [], "Error Count": []}
            for sheet_name, findings_list in findings.items():
                summary_data["Sheet Name"].append(sheet_name)
                summary_data["Error Count"].append(len(findings_list))

            # Écrire les données récapitulatives dans la feuille
            for col, header in enumerate(["Sheet Name", "Error Count"], 1):
                summary_sheet.cell(row=1, column=col, value=header).font = header_font
                summary_sheet.cell(row=1, column=col).fill = header_fill
                summary_sheet.cell(row=1, column=col).border = thin_border

            for row, sheet_name in enumerate(summary_data["Sheet Name"], 2):
                summary_sheet.cell(row=row, column=1, value=sheet_name)
                summary_sheet.cell(row=row, column=2, value=summary_data["Error Count"][row-2])
                summary_sheet.cell(row=row, column=1).border = thin_border
                summary_sheet.cell(row=row, column=2).border = thin_border

            # Ajuster la largeur des colonnes et geler l'en-tête
            for col in summary_sheet.columns:
                max_length = 0
                column = col[0].column_letter  # Récupérer la lettre de la colonne
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                summary_sheet.column_dimensions[column].width = adjusted_width
            summary_sheet.freeze_panes = "A2"

            # Créer le graphique camembert
            pie = PieChart()
            labels = Reference(summary_sheet, min_col=1, min_row=2, max_row=len(summary_data["Sheet Name"]) + 1)
            data = Reference(summary_sheet, min_col=2, min_row=1, max_row=len(summary_data["Sheet Name"]) + 1)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "Summary of Errors by Sheet"
            summary_sheet.add_chart(pie, "E5")

            workbook.save(file_path)
            print(f"Audit completed. Results saved in '{file_path}'.")

    # Demander à l'utilisateur où il veut enregistrer le fichier Excel
    save_excel_file()
else:
    print("No file selected.")
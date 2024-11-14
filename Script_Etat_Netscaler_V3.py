import requests
import ssl
from requests.adapters import HTTPAdapter
from requests.packages.urllib3.util.ssl_ import create_urllib3_context
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from tqdm import tqdm
import getpass
import os
from pwinput import pwinput

# Liste des ciphers pour SSL/TLS
CIPHERS = (
    'ECDH+AESGCM:DH+AESGCM:ECDH+AES256:DH+AES256:ECDH+AES128:DH+AES:ECDH+HIGH:'
    'DH+HIGH:AES256-SHA256:RSA+AESGCM:RSA+AES:RSA+HIGH!aNULL:!eNULL:!MD5'
)


# Désactiver la vérification SSL/TLS
class WeakAdapter(HTTPAdapter):
    def init_poolmanager(self, *args, **kwargs):
        context = create_urllib3_context(ciphers=CIPHERS)
        context.options |= ssl.CERT_NONE
        context.check_hostname = False
        kwargs['ssl_context'] = context
        return super(WeakAdapter, self).init_poolmanager(*args, **kwargs)


try:
    print("____________________________________________________________________\n")
    print("Script Etat NetScaler by Thomas Pruvost\n")
    print("____________________________________________________________________\n")
    ns_adr = input("Entrez l'adresse du NetScaler : ")
    username = input("Entrez le nom d'utilisateur (adm-xxx) : ")
    password = pwinput("Quel est le mot de passe : ", "*")

    baseurl = f'https://{ns_adr}/'
    headers = {'X-NITRO-User': username, 'X-NITRO-PASS': password}
    s = requests.Session()
    s.mount(baseurl, WeakAdapter())

    # Récupération des données avec une barre de progression
    with tqdm(total=3, desc="Récupération des données NetScaler", unit="requêtes") as progress:
        lbvserver_response = s.get(f'{baseurl}nitro/v1/config/lbvserver/', headers=headers, verify=False)
        progress.update(1)
        csvserver_response = s.get(f'{baseurl}nitro/v1/config/csvserver/', headers=headers, verify=False)
        progress.update(1)
        services_response = s.get(f'{baseurl}nitro/v1/config/service/', headers=headers, verify=False)
        progress.update(1)

    # Initialiser le classeur Excel
    wb = Workbook()

    # --- Virtual Servers Sheet ---
    ws = wb.active
    ws.title = 'VSLBs Servers'
    ws.append(['Name', 'Current State', 'Effective State'])
    for cell in ws[1]:
        cell.font = Font(bold=True, underline="single")
        cell.fill = PatternFill(start_color="00C0C0C0", fill_type="solid")

    for item in lbvserver_response.json().get('lbvserver', []):
        ws.append([item.get('name', ''), item.get('curstate', ''), item.get('effectivestate', '')])

    # Mise en forme conditionnelle pour le champ Current State
    for row in range(2, ws.max_row + 1):
        current_state = ws.cell(row=row, column=2).value
        if current_state == 'UP':
            ws.cell(row=row, column=2).fill = PatternFill(start_color="2ecc71", fill_type="solid")  # Vert
        elif current_state == 'DOWN':
            ws.cell(row=row, column=2).fill = PatternFill(start_color="e74c3c", fill_type="solid")  # Rouge
        elif current_state == 'OUT OF SERVICE':
            ws.cell(row=row, column=2).fill = PatternFill(start_color="f1c40f", fill_type="solid")  # Jaune

    # Mise en forme conditionnelle pour le champ Effective State
    for row in range(2, ws.max_row + 1):
        effective_state = ws.cell(row=row, column=3).value
        if effective_state == 'UP':
            ws.cell(row=row, column=3).fill = PatternFill(start_color="2ecc71", fill_type="solid")  # Vert
        elif effective_state == 'DOWN':
            ws.cell(row=row, column=3).fill = PatternFill(start_color="e74c3c", fill_type="solid")  # Rouge
        elif effective_state == 'OUT OF SERVICE':
            ws.cell(row=row, column=3).fill = PatternFill(start_color="f1c40f", fill_type="solid")  # Jaune

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # --- CSV Servers Sheet ---
    ws2 = wb.create_sheet(title='CSV Servers')
    ws2.append(['Name', 'Current State'])
    for cell in ws2[1]:
        cell.font = Font(bold=True, underline="single")
        cell.fill = PatternFill(start_color="00C0C0C0", fill_type="solid")

    for item in csvserver_response.json().get('csvserver', []):
        ws2.append([item.get('name', ''), item.get('curstate', '')])

    # Mise en forme conditionnelle pour le champ Current State
    for row in range(2, ws2.max_row + 1):
        current_state = ws2.cell(row=row, column=2).value
        if current_state == 'UP':
            ws2.cell(row=row, column=2).fill = PatternFill(start_color="2ecc71", fill_type="solid")  # Vert
        elif current_state == 'DOWN':
            ws2.cell(row=row, column=2).fill = PatternFill(start_color="e74c3c", fill_type="solid")  # Rouge
        elif current_state == 'OUT OF SERVICE':
            ws2.cell(row=row, column=2).fill = PatternFill(start_color="f1c40f", fill_type="solid")  # Jaune

    for col in ws2.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # --- Services Sheet ---
    ws3 = wb.create_sheet(title='Services')
    ws3.append(['Name', 'Server State', 'IP Address', 'Port'])
    for cell in ws3[1]:
        cell.font = Font(bold=True, underline="single")
        cell.fill = PatternFill(start_color="00C0C0C0", fill_type="solid")

    for item in services_response.json().get('service', []):
        ws3.append([item.get('name'), item.get('svrstate'), item.get('ipaddress'), item.get('port')])

    # Mise en forme conditionnelle pour le champ Server State
    for row in range(2, ws3.max_row + 1):
        server_state = ws3.cell(row=row, column=2).value
        if server_state == 'UP':
            ws3.cell(row=row, column=2).fill = PatternFill(start_color="2ecc71", fill_type="solid")  # Vert
        elif server_state == 'DOWN':
            ws3.cell(row=row, column=2).fill = PatternFill(start_color="e74c3c", fill_type="solid")  # Rouge
        elif server_state == 'OUT OF SERVICE':
            ws3.cell(row=row, column=2).fill = PatternFill(start_color="f1c40f", fill_type="solid")  # Jaune

    for col in ws3.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws3.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # --- Content Switching Bindings Sheet ---
    ws4 = wb.create_sheet(title='Content Switching Bindings')
    ws4.append(['CSV Server', 'Target VSLB Server', 'Priority'])
    for cell in ws4[1]:
        cell.font = Font(bold=True, underline="single")
        cell.fill = PatternFill(start_color="00C0C0C0", fill_type="solid")

    # Récupérer les associations
    with tqdm(total=len(csvserver_response.json().get('csvserver', [])), desc="Récupération des associations",
              unit="requêtes") as progress:
        for csv_item in csvserver_response.json().get('csvserver', []):
            csv_name = csv_item.get('name')
            csv_binding_url = f'{baseurl}nitro/v1/config/csvserver_binding/{csv_name}'
            csv_binding_response = s.get(csv_binding_url, headers=headers, verify=False)
            progress.update(1)

            if 'csvserver_binding' in csv_binding_response.json():
                for csv_binding in csv_binding_response.json().get('csvserver_binding', []):
                    for policy in csv_binding.get('csvserver_cspolicy_binding', []):
                        target_vlb = policy.get('targetlbvserver', "No Target VSLB Server")
                        priority = policy.get('priority', "No Priority")
                        ws4.append([csv_name, target_vlb, priority])

    for col in ws4.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws4.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # --- Services Bindings Sheet ---
    ws5 = wb.create_sheet(title='Services Bindings')
    ws5.append(['Service Name', 'VLB Server'])
    for cell in ws5[1]:
        cell.font = Font(bold=True, underline="single")
        cell.fill = PatternFill(start_color="00C0C0C0", fill_type="solid")

    # Set to track unique service and VLB combinations
    unique_combinations = {}

    # Récupérer les associations
    with tqdm(total=len(services_response.json().get('service', [])), desc="Récupération des associations de service",
              unit="requêtes") as progress:
        for service in services_response.json().get('service', []):
            service_name = service.get('name')
            service_binding_url = f'{baseurl}nitro/v1/config/svcbindings/{service_name}?view=summary'
            service_binding_response = s.get(service_binding_url, headers=headers, verify=False)
            progress.update(1)

            # Ajouter le service à unique_combinations s'il n'a pas encore été ajouté
            if service_name not in unique_combinations:
                unique_combinations[service_name] = None  # Placeholder for future binding checks

            # Si le service a des bindings, mettre à jour l'association
            if 'svcbindings' in service_binding_response.json():
                for binding in service_binding_response.json().get('svcbindings', []):
                    vlb_name = binding.get('vservername', "No VLB Server")

                    # Update or add only if it's a valid VLB association
                    if vlb_name != "No VLB Server":
                        unique_combinations[service_name] = vlb_name
                    elif unique_combinations[service_name] is None:
                        unique_combinations[service_name] = "No VLB Server"

    # Append only unique combinations to the sheet
    for service_name, vlb_name in unique_combinations.items():
        ws5.append([service_name, vlb_name])

    # Ajustement automatique des colonnes
    for col in ws5.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws5.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # --- Summary Sheet ---
    ws6 = wb.create_sheet(title='Résumé')
    ws6.append(['Type', 'Count'])
    ws6.append(['VSLB Servers', len(lbvserver_response.json().get('lbvserver', []))])
    ws6.append(['CSV Servers', len(csvserver_response.json().get('csvserver', []))])
    ws6.append(['Services', len(services_response.json().get('service', []))])

    # Création du graphique à barres
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Nombre de Serveurs et Services"
    chart.x_axis.title = "Type"
    chart.y_axis.title = "Nombre"
    data = Reference(ws6, min_col=2, min_row=1, max_col=2, max_row=4)
    categories = Reference(ws6, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    ws6.add_chart(chart, "A1")

    # Sauvegarder le fichier Excel
    excel_file_path = f"Etat_Netscaler_{ns_adr}.xlsx"
    wb.save(excel_file_path)
    print(f"Le fichier Excel {excel_file_path} a bien été sauvegardé.")
    os.startfile(excel_file_path)

except requests.exceptions.RequestException as e:
    print(f"Erreur lors de la requête : {e}")

input("Appuyez sur Entrée pour quitter...")
